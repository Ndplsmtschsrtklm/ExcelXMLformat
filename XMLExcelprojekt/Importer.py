import pandas as pd
import xml.etree.ElementTree as eT
from openpyxl import load_workbook
import os
import math
from win10toast import ToastNotifier  # Für Benachrichtigungen


SS_NAMESPACE = "urn:schemas-microsoft-com:office:spreadsheet"
namespaces = {'ss': SS_NAMESPACE}

toaster = ToastNotifier()




# Datei-Pfade
excel_file = "Artikel infos.xlsx"
original_preisliste_file = "Standard/DE_Preislisten.xml"
original_services_file = "Standard/DE_Services.xml"
original_servicebewertungsdaten_file = "Standard/DE_Servicebewertungsdaten.xml"
neue_preisliste_file = r"C:\Users\Ijed.Hadji\Dokumente\SAP Artikel Import2\Standard\NeueListenpreise_US_DE.xlsx"




# Excel-Datei mit openpyxl laden
wb = load_workbook(excel_file)
info_sheet = wb["Infos"]

# Wert aus Zeile 8, Spalte 2 auslesen (für den neuen Dateinamen)
custom_name = info_sheet.cell(row=8, column=2).value

# Einstellungen aus der Excel-Datei lesen
custom_name = info_sheet.cell(row=8, column=2).value
aktivieren_verkauf = info_sheet.cell(row=1, column=5).value == "Ja"
aktivieren_einkauf = info_sheet.cell(row=2, column=5).value == "Ja"

# Überprüfen, ob ein Name vorhanden ist, und Namen zuweisen
if custom_name and isinstance(custom_name, str) and custom_name.strip():
    preisliste_file = f"Standard/DE_Preislisten_{custom_name.strip()}.xml"
    services_file = f"Standard/DE_Services_{custom_name.strip()}.xml"
    servicebewertungsdaten_file = f"Standard/DE_Servicebewertungsdaten_{custom_name.strip()}.xml"
    
    neue_preisliste_file_neu = neue_preisliste_file.replace(".xlsx", f"_{custom_name.strip()}.xlsx")
else:
    # Kein Name angegeben, verwende die ursprünglichen Dateinamen
    preisliste_file = original_preisliste_file
    services_file = original_services_file
    servicebewertungsdaten_file = original_servicebewertungsdaten_file
    neue_preisliste_file_neu = neue_preisliste_file
    
    print("Kein gültiger Name gefunden. Bearbeite die ursprünglichen Dateien.")


# Wert aus Zelle H1 (erste Zeile, achte Spalte) direkt auslesen
produktnummer_start = int(info_sheet["H1"].value)
service_nummer_start = produktnummer_start

# Werte für gültig ab und gültig bis aus der Infos-Tabelle lesen, nur wenn vorhanden
gueltig_ab_cell = info_sheet.cell(row=2, column=11).value
gueltig_bis_cell = info_sheet.cell(row=3, column=11).value

gueltig_ab = gueltig_ab_cell.strftime("%d.%m.%Y") if gueltig_ab_cell else None
gueltig_bis = gueltig_bis_cell.strftime("%d.%m.%Y") if gueltig_bis_cell else None

# Unternehmensnummern dynamisch ermitteln (alle mit "Ja")
unternehmensnummern = []
for row in info_sheet.iter_rows(min_row=2, max_row=info_sheet.max_row, values_only=True):
    if row[1] and row[1].strip().lower() == "ja":
        unternehmensnummern.append(int(row[0]) if isinstance(row[0], float) else row[0])

if not unternehmensnummern:
    raise ValueError("Keine gültigen Unternehmensnummern gefunden! Bitte prüfen Sie die 'Ja'-Markierungen im Infos-Sheet.")

# Pandas DataFrame für die Services-Daten
data = pd.read_excel(excel_file, sheet_name="Services").fillna("")

if os.path.exists(neue_preisliste_file):
    # Lade die Excel-Datei und prüfe, ob das Blatt "Table" existiert
    wb_preisliste = load_workbook(neue_preisliste_file)

    if "Table" in wb_preisliste.sheetnames:
        sheet_preisliste = wb_preisliste["Table"]

        # Lade die Daten aus Artikel infos.xlsx (Sheet "Services")
        df_artikel = pd.read_excel(excel_file, sheet_name="Services")

        if all(col in df_artikel.columns for col in ["lieferantennummer", "Preis", "Mengeneinheit"]):
            lieferantennummern = df_artikel["lieferantennummer"].tolist()
            preise = df_artikel["Preis"].tolist()
            mengeneinheiten = df_artikel["Mengeneinheit"].tolist()  # Mengeneinheit aus Artikel infos

            # **Servicenummer wird von `service_nummer_start` genommen**
            service_num = service_nummer_start

            # Werte ab Zeile 6 einfügen (row_index fängt bei 6 an)
            for row_index, (lieferantennr, preis, menge) in enumerate(zip(lieferantennummern, preise, mengeneinheiten), start=6):
                sheet_preisliste.cell(row=row_index, column=6, value=service_num)  # Spalte 2 = Product ID (Servicenummer)
                sheet_preisliste.cell(row=row_index, column=5, value=lieferantennr)  # Spalte 5 = Lieferantennummer
                
                # **Preisprüfung & Abrundung auf 2 Nachkommastellen**
                if isinstance(preis, float) and len(str(preis).split(".")[1]) > 2:
                    # Benachrichtigung anzeigen
                    toaster.show_toast(
                        "Ungültiger Preis",
                        f"Der Preis {preis} hat mehr als 2 Nachkommastellen und wird auf 2 Nachkommastellen abgerundet.",
                        duration=5
                    )
                    preis = math.floor(preis * 100) / 100  # Abrunden auf 2 Nachkommastellen

                sheet_preisliste.cell(row=row_index, column=10, value=preis)  # Spalte 10 = Preis
                sheet_preisliste.cell(row=row_index, column=11, value="EUR")   # Spalte 11 = Währung EUR
                sheet_preisliste.cell(row=row_index, column=12, value=1)       # Spalte 12 = Preiseinheit 1
                
                # **Mengeneinheit (Spalte 13)**
                menge_cleaned = "EA" if menge == "Each" else menge  # Falls "Each", ersetze mit "EA"
                sheet_preisliste.cell(row=row_index, column=13, value=menge_cleaned)  # Spalte 13 = Mengeneinheit

                service_num += 1  # **Servicenummer wird hochgezählt!**

            # Speichern unter dem neuen Namen
            wb_preisliste.save(neue_preisliste_file_neu)


    
   
# Kontenfindungsgruppe aus der sechsten Spalte in der Tabelle 'Services' auslesen
kontenfindungsgruppen = [int(x) if isinstance(x, float) else x for x in data.iloc[:, 5].tolist()]

# Originaldateien laden
preisliste_tree = eT.parse(original_preisliste_file)
preisliste_root = preisliste_tree.getroot()
services_tree = eT.parse(original_services_file)
services_root = services_tree.getroot()
servicebewertungsdaten_tree = eT.parse(original_servicebewertungsdaten_file)
servicebewertungsdaten_root = servicebewertungsdaten_tree.getroot()


# Setzen des Namespace auf Wurzelebene
preisliste_root.set("xmlns:ss", SS_NAMESPACE)
services_root.set("xmlns:ss", SS_NAMESPACE)
servicebewertungsdaten_root.set("xmlns:ss", SS_NAMESPACE)

def create_cell(index, value, data_type):
    cell = eT.Element(f"{{{SS_NAMESPACE}}}Cell", {'ss:Index': str(index)})
    data = eT.SubElement(cell, f"{{{SS_NAMESPACE}}}Data", {'ss:Type': data_type})
    data.text = str(value)
    return cell

def add_row_to_table(table, values):
    row_element = eT.SubElement(table, f"{{{SS_NAMESPACE}}}Row")
    for index, (value, data_type) in enumerate(values, start=1):
        row_element.append(create_cell(index, value, data_type))


def clear_table_rows_from_index(table, start_index):
    rows = table.findall(f".//{{{SS_NAMESPACE}}}Row")
    for row in rows[start_index - 1:]:  # Entferne alle Zeilen ab dem angegebenen Index (start_index)
        table.remove(row)

# Tabellen aus XML finden
preisliste_table = preisliste_root.find(".//ss:Worksheet[@ss:Name='Positionen']/ss:Table", namespaces)
general_table = preisliste_root.find(".//ss:Worksheet[@ss:Name='Allgemein']/ss:Table", namespaces)
services_table = services_root.find(".//ss:Worksheet[@ss:Name='Allgemein']/ss:Table", namespaces)
detailed_descriptions_table = services_root.find(".//ss:Worksheet[@ss:Name='Detaillierte Beschreibungen']/ss:Table", namespaces)
mengenumrechnungen_table = services_root.find(".//ss:Worksheet[@ss:Name='Mengenumrechnungen']/ss:Table", namespaces)
bewertung_table = services_root.find(".//ss:Worksheet[@ss:Name='Bewertung']/ss:Table", namespaces)
verkaufsorganisationen_table = services_root.find(".//ss:Worksheet[@ss:Name='Verkaufsorganisationen']/ss:Table", namespaces)
verkaufsnotizen_table = services_root.find(".//ss:Worksheet[@ss:Name='Verkaufsnotizen']/ss:Table", namespaces)
einkauf_table = services_root.find(".//ss:Worksheet[@ss:Name='Einkauf']/ss:Table", namespaces)
einkaufsnotizen_table = services_root.find(".//ss:Worksheet[@ss:Name='Einkaufsnotizen']/ss:Table", namespaces)
kundenservicenummern_table = services_root.find(".//ss:Worksheet[@ss:Name='Kundenservicenummern']/ss:Table", namespaces)
lieferantenservicenummern_table = services_root.find(".//ss:Worksheet[@ss:Name='Lieferantenservicenummern']/ss:Table", namespaces)
finanzdaten_allgemein_table = servicebewertungsdaten_root.find(".//ss:Worksheet[@ss:Name='Finanzdaten - allgemein']/ss:Table", namespaces)


# Alte Zeilen in den Services-Tabellen löschen
clear_table_rows_from_index(services_table, 8)
clear_table_rows_from_index(detailed_descriptions_table, 8)
clear_table_rows_from_index(mengenumrechnungen_table, 8)
clear_table_rows_from_index(bewertung_table, 8)
clear_table_rows_from_index(verkaufsorganisationen_table, 8)
clear_table_rows_from_index(verkaufsnotizen_table, 8)
clear_table_rows_from_index(einkauf_table, 8)
clear_table_rows_from_index(einkaufsnotizen_table, 8)
clear_table_rows_from_index(kundenservicenummern_table, 8)
clear_table_rows_from_index(lieferantenservicenummern_table, 8)
clear_table_rows_from_index(finanzdaten_allgemein_table, 8)
clear_table_rows_from_index(general_table, 8)
clear_table_rows_from_index(preisliste_table, 8)

# Sicherstellen, dass alle Tabellen vorhanden sind
if None in [preisliste_table, general_table, services_table, detailed_descriptions_table, mengenumrechnungen_table, bewertung_table, verkaufsorganisationen_table, verkaufsnotizen_table, einkauf_table, einkaufsnotizen_table, kundenservicenummern_table, lieferantenservicenummern_table, finanzdaten_allgemein_table]:
    raise ValueError("Eine der Tabellen wurde nicht gefunden!")

# Separate Nummern für dynamische und statische Tabellen
service_nummer_static = service_nummer_start

# Dynamische Tabellen (basierend auf Unternehmensnummern)
for unternehmensnummer in unternehmensnummern:
    service_nummer_dynamisch = service_nummer_start
    for _, row in data.iterrows():
        if aktivieren_verkauf:
            # Verkaufsorganisationen
            add_row_to_table(verkaufsorganisationen_table, [
                (service_nummer_dynamisch, "Number"),
                (unternehmensnummer + 1, "Number"),
                ("Direktvertrieb", "String"),
                (row['Mengeneinheit'], "String"),
                ("Seco", "String"),
                (" ", "String"),
                (" ", "String"),
                (" ", "String"),
                ("Aktiv", "String")
            ])
            # Verkaufsnotizen
            add_row_to_table(verkaufsnotizen_table, [
                (service_nummer_dynamisch, "Number"),
                (unternehmensnummer + 1, "Number"),
                ("Direktvertrieb", "String"),
                ("Deutsch", "String"),
                (row['Verkaufsnotizen'], "String")
            ])

        
           

        # Gemeinsame Tabellen (immer ausfüllen)
        add_row_to_table(finanzdaten_allgemein_table, [
            (service_nummer_dynamisch, "Number"),
            (unternehmensnummer, "Number"),
            (kontenfindungsgruppen[0], "String")
        ])
        # Bewertung
        add_row_to_table(bewertung_table, [
            (service_nummer_dynamisch, "Number"),
            (unternehmensnummer, "Number")
        ])
        service_nummer_dynamisch += 1

# Statische Tabellen (einmalig füllen)
if aktivieren_verkauf:
    # Allgemein
    add_row_to_table(general_table, [
        ("Ijed", "String"),
        (gueltig_ab if gueltig_ab else " ", "String"),
        (gueltig_bis if gueltig_bis else " ", "String")
    ])
    # Positionen
    for _, row in data.iterrows():
        preis = row['Preis']
        if isinstance(preis, float) and len(str(preis).split(".")[1]) > 2:
            # Benachrichtigung anzeigen
            toaster.show_toast(
                "Ungültiger Preis",
                f"Der Preis {preis} hat mehr als 2 Nachkommastellen und wird auf 2 Nachkommastellen abgerundet.",
                duration=5
            )
            # Abrunden auf 2 Nachkommastellen
            preis = math.floor(preis * 100) / 100

        # Verwende den abgerundeten Preis in der Tabelle
        add_row_to_table(preisliste_table, [
            ("Ijed", "String"),
            (service_nummer_start, "Number"),
            ("2", "String"),
            (row['Produktkategorie'], "String"),
            (preis, "Number"),  # Hier wird der abgerundete Preis verwendet
            (" ", "String"),
            (row['Mengeneinheit'], "String")
        ])
        service_nummer_start += 1
   
       

for _, row in data.iterrows():
    if aktivieren_verkauf:
        # Kundenservicenummern (nur bei aktivieren_verkauf)
        add_row_to_table(kundenservicenummern_table, [
            (service_nummer_static, "Number"),
            (row['kundennummer'], "String"),
            (row['Kundenservicenummer'], "String")
        ])
        

    if aktivieren_einkauf:
        # Lieferantenservicenummern (nur bei aktivieren_einkauf)
        add_row_to_table(lieferantenservicenummern_table, [
            (service_nummer_static, "Number"),
            (row['lieferantennummer'], "String"),
            (row['liefarantenservicenummer'], "String")
        ])

         # Einkauf
        add_row_to_table(einkauf_table, [
            (service_nummer_static, "Number"),
            ("Aktiv", "String"),
            (row['Mengeneinheit'], "String")
            ])
            # Einkaufsnotizen
        add_row_to_table(einkaufsnotizen_table, [
            (service_nummer_static, "Number"),
            ("Deutsch", "String"),
            (row['Einkaufnotizen'], "String")
            ])

    
   

    # Services
    add_row_to_table(services_table, [
        (service_nummer_static, "Number"),
        ("Deutsch", "String"),
        (row['Produktbezeichnung'], "String"),
        (row['Produktkategorie'], "String"),
        (" ", "String"),
        (row['Mengeneinheit'], "String")
    ])

    # Detaillierte Beschreibungen
    add_row_to_table(detailed_descriptions_table, [
        (service_nummer_static, "Number"),
        ("Deutsch", "String"),
        (row['detaillierte beschreibung'], "String")
    ])

    # Mengenumrechnungen
    fifth_value = "HUR" if row['Mengeneinheit'] not in ["HUR"] else "Each"
    add_row_to_table(mengenumrechnungen_table, [
        (service_nummer_static, "Number"),
        (1, "Number"),
        (row['Mengeneinheit'], "String"),
        (1, "Number"),
        (fifth_value, "String")
    ])
    service_nummer_static += 1

# `ExpandedRowCount` aktualisieren
for table in [preisliste_table, general_table, services_table, detailed_descriptions_table, mengenumrechnungen_table, bewertung_table, verkaufsorganisationen_table, verkaufsnotizen_table, einkauf_table, einkaufsnotizen_table, kundenservicenummern_table, lieferantenservicenummern_table, finanzdaten_allgemein_table]:
    table.set(f"{{{SS_NAMESPACE}}}ExpandedRowCount", str(len(table.findall(f'.//{{{SS_NAMESPACE}}}Row'))))

# Dateien speichern mit neuen Namen
preisliste_tree.write(preisliste_file, encoding='utf-8', xml_declaration=True)
services_tree.write(services_file, encoding='utf-8', xml_declaration=True)
servicebewertungsdaten_tree.write(servicebewertungsdaten_file, encoding='utf-8', xml_declaration=True)

print(f"Dateien erfolgreich aktualisiert und umbenannt:")
print(f"- {preisliste_file}")
print(f"- {services_file}")
print(f"- {servicebewertungsdaten_file}")
